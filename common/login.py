#!/usr/bin/env python
# _*_ coding:utf-8 _*_
__author__ = 'BIN'
import os,sys
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
import requests,unittest,time
from common.URL import *
from openpyxl import load_workbook
from common.config import *
import shutil
class Login():
    "登录类"
    def adminlogin(self):
        "这是运营端登陆"
        headers = \
        {
            "Content-Type": "application/x-www-form-urlencoded"
        }
        body = \
        {
            "username": "ADMin",
            "password": "123456.",
            "verificationCode": "99999",
            "loginStatus": "0"
        }
        r = requests.post(adminurl("login"), headers=headers, data=body)
        token = (r.json()["data"]["token"])
        if not os.path.exists(TEST_TOKEN):  # 检查文件是否存在
            shutil.copyfile(TEST_CONFIG,TEST_TOKEN) #不存在则从测试用例复制一份
        self.wb = load_workbook(TEST_TOKEN)
        self.ws = self.wb.active
        self.max_row = self.ws.max_row  #获取最大的有效行数
        for i in range(1,self.max_row):
            self.ws.cell(i+1, 5, '{"Authorization":"' + token + '"}')#把获取的token写进excel表格的每行用例中
            self.wb.save(TEST_TOKEN)#保存表格
    def merchantlogin(self):
        "这是商户端登录"
        headers = \
        {
            "Content-Type": "application/x-www-form-urlencoded"
        }
        body = \
        {
            "username": "LMKJ",
            "password": "123456.",
            "verificationCode": "99999",
            "loginStatus": "1"
        }
        r = requests.post(merchanturl("login"), headers=headers, data=body)
        token = (r.json()["data"]["token"])
        return token

# if __name__ == '__main__':
#     Login().adminlogin()
