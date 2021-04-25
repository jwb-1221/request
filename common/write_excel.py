#!/usr/bin/env python
# _*_ coding:utf-8 _*_
__author__ = 'BIN'

import os,sys
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from common import config
#导入复制文件模块
import shutil
from openpyxl import load_workbook
#styles:样式，font：字体，Alignment：对齐 阵营 对准
from openpyxl.styles import Font,Alignment,colors
import configparser as cparser
# --------- 读取config.ini配置文件 ---------------
cf = cparser.ConfigParser()
cf.read(config.CONFIG,encoding='utf-8')
name = cf.get("tester","name")
font_RED = Font(name='宋体', color='FF0000', bold=True)
font_GREEN = Font(name='宋体', color='00ff00', bold=True)
font_purple = Font(name='宋体', color='9900cc', bold=True)
# align = Alignment(horizontal='center', vertical='center')

class WriteExcel():
    """文件写入数据"""
    def __init__(self,fileName):
        """初始化"""
        self.filename = fileName
        if not os.path.exists(self.filename):
            # 文件不存在，则拷贝模板文件至指定报告目录下
            shutil.copyfile(config.TEST_TOKEN,config.TEST_RESULT)
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def save_excel(self):
        """保存文件"""
        self.wb.save(self.filename)

    def write_result_pass(self,row_n,value="pass"):
        """写入测试成功结果"""
        self.ws.cell(row_n, 8, value).font = font_GREEN
        WriteExcel.save_excel(self)

    def write_result_fail(self,row_n,value="fail"):
        """写入测试失败结果"""
        self.ws.cell(row_n, 8, value).font = font_RED
        WriteExcel.save_excel(self)

    def write_anomaly(self,row_n,value= None):
        """如果异常写入值方法"""
        self.ws.cell(row_n,9, value).font = font_purple#写接口返回信息
        WriteExcel.save_excel(self)

    def write_name(self,row_n,value = name):
        """写入测试员方法"""
        self.ws.cell(row_n,10, value).font = font_purple#写入测试员
        WriteExcel.save_excel(self)




# if __name__=='__main__':
#     WriteExcel(config.TEST_RESULT).write_data(5,'FAIL')
